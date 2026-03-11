"""
置換モジュール
Excelファイル内のセル値を検索・置換し、上書き保存する。
.xlsx のみ書き込み対応（.xls は読み取り専用ライブラリのため非対応）。
"""
import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional


@dataclass
class ReplaceRecord:
    """1件の置換記録"""
    file_path: str
    sheet_name: str
    cell_address: str
    row: int
    col: int
    matched_keyword: str
    before: str
    after: str


@dataclass
class ReplaceResult:
    """ファイル単位の置換結果"""
    file_path: str
    records: List[ReplaceRecord] = field(default_factory=list)
    error: Optional[str] = None
    skipped: bool = False
    dry_run: bool = False
    backed_up: bool = False

    @property
    def replace_count(self) -> int:
        return len(self.records)


def _col_to_letter(col: int) -> str:
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _cell_address(row: int, col: int) -> str:
    return f"{_col_to_letter(col)}{row}"


def replace_in_xlsx(
    file_path: Path,
    replace_map: Dict[str, str],
    use_regex: bool = False,
    backup: bool = True,
    dry_run: bool = False,
) -> ReplaceResult:
    """
    .xlsx ファイル内のセルを検索・置換する。

    Args:
        file_path: 対象ファイルパス
        replace_map: {検索キーワード: 置換文字列} の辞書
        use_regex: 正規表現使用フラグ
        backup: バックアップを作成するか（.bak ファイル）
        dry_run: Trueの場合、実際には書き込まない（プレビューのみ）

    Returns:
        ReplaceResult
    """
    try:
        import openpyxl
    except ImportError:
        return ReplaceResult(file_path=str(file_path), error="openpyxlがインストールされていません", skipped=True)

    result = ReplaceResult(file_path=str(file_path), dry_run=dry_run)

    # パターンコンパイル
    patterns = {}
    for kw, replacement in replace_map.items():
        if use_regex:
            try:
                patterns[kw] = (re.compile(kw, re.IGNORECASE), replacement)
            except re.error as e:
                result.error = f"不正な正規表現 '{kw}': {e}"
                result.skipped = True
                return result
        else:
            patterns[kw] = (None, replacement)

    try:
        # read_only=False で開く（書き込み対応）
        wb = openpyxl.load_workbook(str(file_path), data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                for cell in row:
                    # 文字列セルのみ対象
                    if cell.value is None or not isinstance(cell.value, str):
                        continue

                    original = cell.value
                    current = original
                    col_idx = cell.column
                    addr = _cell_address(row_idx, col_idx)

                    for kw, (pattern, replacement) in patterns.items():
                        if use_regex:
                            if pattern.search(current):
                                new_val = pattern.sub(replacement, current)
                                result.records.append(ReplaceRecord(
                                    file_path=str(file_path),
                                    sheet_name=sheet_name,
                                    cell_address=addr,
                                    row=row_idx,
                                    col=col_idx,
                                    matched_keyword=kw,
                                    before=current,
                                    after=new_val,
                                ))
                                current = new_val
                        else:
                            if kw.lower() in current.lower():
                                # 大文字小文字を保持した置換
                                new_val = re.sub(re.escape(kw), replacement, current, flags=re.IGNORECASE)
                                result.records.append(ReplaceRecord(
                                    file_path=str(file_path),
                                    sheet_name=sheet_name,
                                    cell_address=addr,
                                    row=row_idx,
                                    col=col_idx,
                                    matched_keyword=kw,
                                    before=original,
                                    after=new_val,
                                ))
                                current = new_val

                    # 実際に書き換え（dry_run でなければ）
                    if not dry_run and current != original:
                        cell.value = current

        if not dry_run and result.records:
            # バックアップ作成
            if backup:
                bak_path = file_path.with_suffix(file_path.suffix + ".bak")
                shutil.copy2(file_path, bak_path)
                result.backed_up = True

            wb.save(str(file_path))

        wb.close()

    except Exception as e:
        result.error = str(e)
        result.skipped = True

    return result


def replace_files(
    file_list: List[Path],
    replace_map: Dict[str, str],
    use_regex: bool = False,
    backup: bool = True,
    dry_run: bool = False,
    logger=None,
    progress_callback=None,
) -> List[ReplaceResult]:
    """
    複数ファイルを順次置換する。

    Args:
        file_list: 対象ファイルリスト
        replace_map: {検索キーワード: 置換文字列}
        use_regex: 正規表現フラグ
        backup: バックアップ作成フラグ
        dry_run: プレビューのみフラグ
        logger: ロガーインスタンス
        progress_callback: 進捗コールバック callback(current, total, file_path, total_replaced)

    Returns:
        ReplaceResult のリスト
    """
    total = len(file_list)
    results = []
    mode_str = "プレビュー" if dry_run else "置換"

    if logger:
        logger.info(f"置換開始: {total}ファイル, dry_run={dry_run}")

    for i, fp in enumerate(file_list, start=1):
        suffix = fp.suffix.lower()
        if suffix == ".xls":
            result = ReplaceResult(
                file_path=str(fp),
                error=".xls 形式は置換非対応です（.xlsx に変換してください）",
                skipped=True,
            )
        else:
            result = replace_in_xlsx(fp, replace_map, use_regex, backup, dry_run)

        results.append(result)

        if result.skipped:
            if logger:
                logger.warning(f"{fp.name}: スキップ ({result.error})")
        else:
            if logger:
                logger.info(f"{fp.name}: {result.replace_count}件{mode_str}")

        if progress_callback:
            total_replaced = sum(r.replace_count for r in results if not r.skipped)
            progress_callback(i, total, str(fp), total_replaced)

    if logger:
        total_replaced = sum(r.replace_count for r in results if not r.skipped)
        errors = sum(1 for r in results if r.skipped)
        logger.info(f"置換完了: 総置換={total_replaced}件, エラー={errors}件")

    return results
