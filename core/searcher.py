"""
検索エンジンモジュール
Excelファイルのセルを正規表現または通常文字列で検索し、マッチ結果を返す。
"""
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, List, Optional

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


@dataclass
class MatchResult:
    """1件のマッチ結果"""
    file_path: str
    sheet_name: str
    cell_address: str    # 例: "A1"
    row: int
    col: int
    matched_keyword: str
    cell_value: str
    use_regex: bool

    def to_dict(self) -> dict:
        return {
            "file_path": self.file_path,
            "sheet_name": self.sheet_name,
            "cell_address": self.cell_address,
            "row": self.row,
            "col": self.col,
            "matched_keyword": self.matched_keyword,
            "cell_value": self.cell_value,
            "search_mode": "正規表現" if self.use_regex else "通常検索",
        }


@dataclass
class SearchResult:
    """ファイル単位の検索結果"""
    file_path: str
    matches: List[MatchResult] = field(default_factory=list)
    error: Optional[str] = None
    skipped: bool = False

    @property
    def match_count(self) -> int:
        return len(self.matches)


def _col_to_letter(col: int) -> str:
    """列番号（1始まり）をExcel列文字に変換。例: 1→A, 26→Z, 27→AA"""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _cell_address(row: int, col: int) -> str:
    """行・列番号（1始まり）からExcelセルアドレスを生成。例: (1,1)→A1"""
    return f"{_col_to_letter(col)}{row}"


def _compile_patterns(keywords: List[str], use_regex: bool) -> List[tuple]:
    """
    キーワードリストをコンパイル済みパターンリストに変換する。

    Returns:
        [(keyword, compiled_pattern | None)] のリスト
        use_regex=Falseの場合、compiled_patternはNone
    """
    patterns = []
    for kw in keywords:
        if use_regex:
            try:
                pattern = re.compile(kw, re.IGNORECASE)
                patterns.append((kw, pattern))
            except re.error as e:
                raise ValueError(f"不正な正規表現パターン '{kw}': {e}")
        else:
            patterns.append((kw, None))
    return patterns


def _search_xlsx(file_path: Path, patterns: List[tuple], use_regex: bool) -> List[MatchResult]:
    """openpyxlで.xlsxファイルを検索する"""
    matches = []
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cell_str = str(cell.value)
                    row_idx = cell.row
                    col_idx = cell.column
                    addr = _cell_address(row_idx, col_idx)

                    for kw, pattern in patterns:
                        if use_regex:
                            if pattern.search(cell_str):
                                matches.append(MatchResult(
                                    file_path=str(file_path),
                                    sheet_name=sheet_name,
                                    cell_address=addr,
                                    row=row_idx,
                                    col=col_idx,
                                    matched_keyword=kw,
                                    cell_value=cell_str,
                                    use_regex=True,
                                ))
                        else:
                            if kw.lower() in cell_str.lower():
                                matches.append(MatchResult(
                                    file_path=str(file_path),
                                    sheet_name=sheet_name,
                                    cell_address=addr,
                                    row=row_idx,
                                    col=col_idx,
                                    matched_keyword=kw,
                                    cell_value=cell_str,
                                    use_regex=False,
                                ))
    finally:
        wb.close()
    return matches


def _search_xls(file_path: Path, patterns: List[tuple], use_regex: bool) -> List[MatchResult]:
    """xlrdで.xlsファイルを検索する"""
    matches = []
    wb = xlrd.open_workbook(str(file_path))
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        sheet_name = ws.name
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                cell_val = ws.cell_value(row_idx, col_idx)
                if cell_val is None or cell_val == "":
                    continue
                cell_str = str(cell_val)
                addr = _cell_address(row_idx + 1, col_idx + 1)

                for kw, pattern in patterns:
                    if use_regex:
                        if pattern.search(cell_str):
                            matches.append(MatchResult(
                                file_path=str(file_path),
                                sheet_name=sheet_name,
                                cell_address=addr,
                                row=row_idx + 1,
                                col=col_idx + 1,
                                matched_keyword=kw,
                                cell_value=cell_str,
                                use_regex=True,
                            ))
                    else:
                        if kw.lower() in cell_str.lower():
                            matches.append(MatchResult(
                                file_path=str(file_path),
                                sheet_name=sheet_name,
                                cell_address=addr,
                                row=row_idx + 1,
                                col=col_idx + 1,
                                matched_keyword=kw,
                                cell_value=cell_str,
                                use_regex=False,
                            ))
    return matches


def search_file(file_path: Path, patterns: List[tuple], use_regex: bool) -> SearchResult:
    """
    単一Excelファイルを検索する。

    Args:
        file_path: 検索対象ファイルパス
        patterns: [(keyword, compiled_pattern)] のリスト
        use_regex: 正規表現フラグ

    Returns:
        SearchResult
    """
    result = SearchResult(file_path=str(file_path))
    try:
        suffix = file_path.suffix.lower()
        if suffix == ".xlsx":
            if not OPENPYXL_AVAILABLE:
                result.error = "openpyxlがインストールされていません"
                result.skipped = True
                return result
            result.matches = _search_xlsx(file_path, patterns, use_regex)
        elif suffix == ".xls":
            if not XLRD_AVAILABLE:
                result.error = "xlrdがインストールされていません"
                result.skipped = True
                return result
            result.matches = _search_xls(file_path, patterns, use_regex)
        else:
            result.error = f"サポートされていないファイル形式: {suffix}"
            result.skipped = True
    except Exception as e:
        result.error = str(e)
        result.skipped = True

    return result


class ExcelSearcher:
    """Excelファイル検索エンジン"""

    def __init__(
        self,
        keywords: List[str],
        use_regex: bool = False,
        max_workers: int = 4,
        logger=None,
        progress_callback: Optional[Callable] = None,
    ):
        """
        Args:
            keywords: 検索キーワードリスト（最大10個）
            use_regex: 正規表現使用フラグ
            max_workers: 並列処理スレッド数
            logger: ロガーインスタンス
            progress_callback: 進捗コールバック callback(current, total, file_path, match_count)
        """
        if len(keywords) > 10:
            raise ValueError(f"キーワードは最大10個までです（指定: {len(keywords)}個）")
        if len(keywords) == 0:
            raise ValueError("キーワードを1個以上指定してください")

        self.keywords = keywords
        self.use_regex = use_regex
        self.max_workers = max_workers
        self.logger = logger
        self.progress_callback = progress_callback

        # パターンコンパイル（Valueエラーを早期検出）
        self.patterns = _compile_patterns(keywords, use_regex)

    def search(self, file_list: List[Path]) -> List[SearchResult]:
        """
        複数ファイルを並列検索する。

        Args:
            file_list: 検索対象ファイルパスリスト

        Returns:
            SearchResult のリスト（ファイル順）
        """
        total = len(file_list)
        results = [None] * total

        if self.logger:
            mode_str = "正規表現" if self.use_regex else "通常検索"
            self.logger.info(f"検索開始: {total}ファイル, モード={mode_str}, キーワード={self.keywords}")

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_idx = {
                executor.submit(search_file, fp, self.patterns, self.use_regex): (i, fp)
                for i, fp in enumerate(file_list)
            }

            completed = 0
            for future in as_completed(future_to_idx):
                i, fp = future_to_idx[future]
                try:
                    result = future.result()
                except Exception as e:
                    result = SearchResult(file_path=str(fp), error=str(e), skipped=True)

                results[i] = result
                completed += 1

                if result.skipped:
                    if self.logger:
                        self.logger.warning(f"読み込みエラー: {fp.name} ({result.error}) - スキップ")
                else:
                    if self.logger:
                        self.logger.info(f"{fp.name}: {result.match_count}件マッチ")

                if self.progress_callback:
                    total_matches = sum(
                        r.match_count for r in results if r is not None and not r.skipped
                    )
                    self.progress_callback(completed, total, str(fp), total_matches)

        if self.logger:
            total_matches = sum(r.match_count for r in results if r and not r.skipped)
            errors = sum(1 for r in results if r and r.skipped)
            self.logger.info(f"検索完了: 総マッチ={total_matches}件, エラー={errors}件")

        return results
